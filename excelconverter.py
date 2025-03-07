import os
import sqlite3
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime, timedelta
import subprocess
import openpyxl

# Global employee list
employee_list = {}
history_window_open = False
history_window = None
current_employee_file = None
current_dat_file = None


# Global variables
history_window_open = False
history_window = None  # Ensure global reference to history window


def create_database():
    conn = sqlite3.connect("conversion_history.db")
    cursor = conn.cursor()

    # Create conversion_history table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS conversions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        filename TEXT,
        converted_at TEXT,
        output_path TEXT
    )
    """)

    # Create employees table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL
        )
    """)

    conn.commit()
    conn.close()

def format_dtr_summary_sheet(writer, df, month_year):
    """
    Formats the DTR summary sheet to match the layout in the example image
    while using the existing data processed by filter_in_out_entries
    """
    from datetime import datetime
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    sheet_name = f"DTR - {month_year}"
    
    # Get the workbook and worksheet directly
    workbook = writer.book
    worksheet = workbook[sheet_name]
    
    # Get all date columns from the existing data
    date_columns = [col for col in df.columns if "Time In" in col or "Time Out" in col]
    unique_dates = sorted(set([col.split(" (")[0] for col in date_columns if " (" in col]))
    
    # Calculate total number of columns used in the sheet
    # 2 columns for ID and NAME + 2 columns for each date (AM/PM)
    total_columns = 2 + (len(unique_dates) * 2)
    
    # Clear existing merges to avoid conflicts
    for merged_range in list(worksheet.merged_cells.ranges):
        worksheet.unmerge_cells(str(merged_range))
    
    # Convert the column number to Excel column letter
    last_column_letter = get_column_letter(total_columns)
    
    # Define orange fill for title
    orange_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
    
    # Apply the title and formatting to A1 BEFORE merging
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = month_year.upper()
    title_cell.font = Font(name='Calibri', size=20, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.fill = orange_fill
    
    # Merge the title across all columns
    worksheet.merge_cells(f'A1:{last_column_letter}1')
    
    # Set the height for title row (row 1)
    worksheet.row_dimensions[1].height = 30
    
    # Set the height for spacer row (row 2)
    worksheet.row_dimensions[2].height = 15
    
    # Starting row and column for the ID and NAME headers
    header_row = 3
    start_col = 3  # Column C (after ID and Name)
    
    # Set heights for header rows (3, 4, 5)
    worksheet.row_dimensions[header_row].height = 25     # Date row
    worksheet.row_dimensions[header_row+1].height = 25   # Day name row
    worksheet.row_dimensions[header_row+2].height = 25   # AM/PM row
    
    # Add ID and NAME headers
    id_cell = worksheet.cell(row=header_row, column=1)
    id_cell.value = "ID"
    id_cell.font = Font(bold=True)
    id_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    name_cell = worksheet.cell(row=header_row, column=2)
    name_cell.value = "NAME"
    name_cell.font = Font(bold=True)
    name_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set width for ID and NAME columns
    worksheet.column_dimensions['A'].width = 15  # For ID column
    worksheet.column_dimensions['B'].width = 30  # For NAME column
    
    # Define border style
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Apply borders to the ID and NAME cells
    id_cell.border = thin_border
    name_cell.border = thin_border
    
    # Merge NAME cell vertically across rows 3, 4, and 5
    worksheet.merge_cells(start_row=header_row, start_column=2, 
                        end_row=header_row+2, end_column=2)
    
    # Merge ID cell vertically across rows 3, 4, and 5
    worksheet.merge_cells(start_row=header_row, start_column=1, 
                        end_row=header_row+2, end_column=1)
    
    # Define color fills based on day of week
    day_fills = {
        'MON': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        'TUE': PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid'),
        'WED': PatternFill(start_color='D9C3E6', end_color='D9C3E6', fill_type='solid'),
        'THU': PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid'),
        'FRI': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
        'SAT': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        'SUN': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    }
    
    # Row colors for employee rows (alternating colors)
    row_colors = [
        PatternFill(start_color='FFD6D6', end_color='FFD6D6', fill_type='solid'),  # Light Red
        PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),  # Light Green
        PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),  # Light Yellow
        PatternFill(start_color='DEEBF7', end_color='DEEBF7', fill_type='solid'),  # Light Blue
    ]
    
    # Add date headers and set column widths for AM/PM columns
    for i, date_str in enumerate(unique_dates):
        date = datetime.strptime(date_str, '%Y-%m-%d')
        # Date column
        date_col = start_col + i*2
        day_name = date.strftime('%a').upper()
        formatted_date = date.strftime('%d/%m/%Y')
        
        # Get the appropriate fill for this day
        current_fill = day_fills.get(day_name, None)
        
        # Set width for AM/PM columns to exactly 12.00 (approx 115 pixels)
        am_col_letter = get_column_letter(date_col)
        pm_col_letter = get_column_letter(date_col+1)
        worksheet.column_dimensions[am_col_letter].width = 12.00
        worksheet.column_dimensions[pm_col_letter].width = 12.00
        
        # Date cell formatting
        date_cell = worksheet.cell(row=header_row, column=date_col)
        date_cell.value = formatted_date
        date_cell.font = Font(bold=True)
        date_cell.alignment = Alignment(horizontal='center')
        date_cell.border = thin_border
        
        next_date_cell = worksheet.cell(row=header_row, column=date_col+1)
        next_date_cell.border = thin_border
        
        # Merge date cells
        worksheet.merge_cells(start_row=header_row, start_column=date_col, 
                             end_row=header_row, end_column=date_col+1)
        
        # Day name cell
        day_row = header_row + 1
        day_cell = worksheet.cell(row=day_row, column=date_col)
        day_cell.value = day_name
        day_cell.font = Font(bold=True)
        day_cell.alignment = Alignment(horizontal='center')
        day_cell.border = thin_border
        
        # Apply fill to day cell
        if current_fill:
            day_cell.fill = current_fill
        
        next_day_cell = worksheet.cell(row=day_row, column=date_col+1)
        next_day_cell.border = thin_border
        if current_fill:
            next_day_cell.fill = current_fill
        
        # Merge day cells
        worksheet.merge_cells(start_row=day_row, start_column=date_col, 
                             end_row=day_row, end_column=date_col+1)
        
        # AM/PM headers
        am_pm_row = header_row + 2
        am_cell = worksheet.cell(row=am_pm_row, column=date_col)
        pm_cell = worksheet.cell(row=am_pm_row, column=date_col+1)
        
        am_cell.value = "AM"
        pm_cell.value = "PM"
        
        # Style AM/PM headers
        am_cell.font = Font(bold=True)
        pm_cell.font = Font(bold=True)
        am_cell.alignment = Alignment(horizontal='center')
        pm_cell.alignment = Alignment(horizontal='center')
        am_cell.border = thin_border
        pm_cell.border = thin_border
        
        # Apply fill to AM/PM cells
        if current_fill:
            am_cell.fill = current_fill
            pm_cell.fill = current_fill
    
    # Employee data row height (approximately 45 pixels = 75.00 points)
    row_height = 45
    
    # Clear any existing data and write employee data starting from row 6
    employee_start_row = header_row + 3  # This is row 6 in the worksheet
    
    # Write employee data
    for i, row in df.iterrows():
        # Set row height for employee rows
        worksheet.row_dimensions[employee_start_row + i].height = row_height
        
        # Write ID and Name
        id_cell = worksheet.cell(row=employee_start_row + i, column=1)
        name_cell = worksheet.cell(row=employee_start_row + i, column=2)
        
        id_cell.value = row.get('Employee No.', '')
        name_cell.value = row.get(df.columns[1], '') if len(df.columns) > 1 else ''
        
        # Format ID and NAME cells
        id_cell.alignment = Alignment(horizontal='center')
        name_cell.alignment = Alignment(horizontal='left')
        id_cell.border = thin_border
        name_cell.border = thin_border
        
        
        # Write time entries
        for date_idx, date_str in enumerate(unique_dates):
            # Find the corresponding columns in the DataFrame
            am_col = next((col for col in df.columns if date_str in col and "Time In" in col), None)
            pm_col = next((col for col in df.columns if date_str in col and "Time Out" in col), None)
            
            if am_col and pm_col:
                am_value = row.get(am_col, '')
                pm_value = row.get(pm_col, '')
                
                # Write values
                am_cell = worksheet.cell(row=employee_start_row + i, column=start_col + date_idx*2)
                pm_cell = worksheet.cell(row=employee_start_row + i, column=start_col + date_idx*2 + 1)
                
            
                
                # Center align the time cells and add borders
                am_cell.alignment = Alignment(horizontal='center', vertical='center')
                pm_cell.alignment = Alignment(horizontal='center', vertical='center')
                am_cell.border = thin_border
                pm_cell.border = thin_border
                
                # Apply conditional formatting here if needed
    
    # Create frozen panes to keep headers and ID/Name columns visible when scrolling
    # Freeze panes at cell C6 (row 6, column 3)
    worksheet.freeze_panes = 'C6'
    
    
def filter_in_out_entries(df):
    # Keep your original function mostly intactpy excelconverter.py
    # Validate input DataFrame
    if df.shape[1] < 2:
        return df  # Return original if insufficient columns

    # Identify columns
    first_col = df.columns[0]  # Employee Name column
    time_col = df.columns[1]  # Timestamp column

    # Convert timestamp and extract date and time
    df[time_col] = pd.to_datetime(df[time_col], errors='coerce')
    df['Date'] = df[time_col].dt.strftime('%Y-%m-%d')
    df['Time'] = df[time_col].dt.strftime('%H:%M:%S')
    df['Hour'] = df[time_col].dt.hour

    # Group logs by employee and date
    grouped = df.groupby([first_col, 'Date'])['Time'].agg(list).reset_index()
    
    # Add Hours column for processing
    grouped['Hours'] = grouped.apply(
        lambda row: df[(df[first_col] == row[first_col]) & (df['Date'] == row['Date'])]['Hour'].tolist(),
        axis=1
    )
    
    # Determine time-in and time-out logs
    def process_daily_logs(times, hours):
        # Sort times to ensure chronological order
        sorted_times = sorted(zip(times, hours), key=lambda x: x[1])
        
        if len(sorted_times) == 1:
            # Single log handling
            time, hour = sorted_times[0]
            if hour < 12:
                return time, "No Out"
            else:
                return "No In", time
        else:
            # Multiple logs
            return sorted_times[0][0], sorted_times[-1][0]

    # Apply log processing
    grouped[['Time In', 'Time Out']] = grouped.apply(
        lambda row: pd.Series(process_daily_logs(row['Time'], row['Hours'])), 
        axis=1
    )
    grouped = grouped.drop(columns=['Time', 'Hours'])

    # Generate full date range for the month
    if not df.empty:
        min_date = pd.to_datetime(df['Date'].min())
        max_date = pd.to_datetime(df['Date'].max())
        
        # Get first and last day of the month
        year, month = min_date.year, min_date.month
        first_day = datetime(year, month, 1)
        if month == 12:
            last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = datetime(year, month + 1, 1) - timedelta(days=1)
        
        all_dates = pd.date_range(first_day, last_day).strftime('%Y-%m-%d')

        # Create full index for all employees and dates
        all_employees = df[first_col].unique()
        full_index = pd.MultiIndex.from_product([all_employees, all_dates], names=[first_col, 'Date'])
        grouped = grouped.set_index([first_col, 'Date']).reindex(full_index).reset_index()

    # Determine day of the week
    grouped['DayOfWeek'] = grouped['Date'].apply(lambda x: pd.to_datetime(x).strftime('%A'))

    # Mark absences and weekend days
    def mark_absences(row):
        if pd.isna(row['Time In']) and pd.isna(row['Time Out']):
            if row['DayOfWeek'] == 'Saturday':
                return "Saturday", "Saturday"
            elif row['DayOfWeek'] == 'Sunday':
                return "Sunday", "Sunday"
            else:
                return "Absent", "Absent"
        return row['Time In'], row['Time Out']

    grouped[['Time In', 'Time Out']] = grouped.apply(mark_absences, axis=1, result_type="expand")

    # Add Employee No. mapping (assuming employee_list is defined elsewhere)
    name_to_id = {v: k for k, v in employee_list.items()}
    grouped['Employee No.'] = grouped[first_col].map(name_to_id)

    # For float columns, fill with 0.0 or another appropriate numeric value
    grouped.fillna({col: 0.0 for col in grouped.select_dtypes(include=['float64']).columns}, inplace=True)
    # For string columns, fill with empty string
    grouped.fillna({col: '' for col in grouped.select_dtypes(include=['object']).columns}, inplace=True)

    # Pivot and format results
    result = grouped.pivot(index=['Employee No.', first_col], columns='Date', values=['Time In', 'Time Out'])
    result = result.swaplevel(axis=1).sort_index(axis=1, level=0)
    result.columns = [f"{date} ({datetime.strptime(date, '%Y-%m-%d').strftime('%A')}) {status}" 
                      for date, status in result.columns]

    return result.reset_index().sort_values(by="Employee No.").reset_index(drop=True)

def convert_batch_to_excel(files):
    for dat_file in files:
        try:
            df = pd.read_csv(dat_file, delimiter="\t", header=None)

            if df.empty:
                messagebox.showerror("Error", f"The file {dat_file} is empty.")
                return

            if df.shape[1] < 2:
                messagebox.showerror("Error", "DAT file must have at least two columns (ID and Timestamp).")
                return

            df.columns = ["Name", "Timestamp"] + [f"Col_{i}" for i in range(2, df.shape[1])]
            df["Name"] = df["Name"].map(lambda x: employee_list.get(int(float(x)) if isinstance(x, (int, float, str)) and str(x).replace('.', '', 1).isdigit() else x, str(x)))

            df["Timestamp"] = pd.to_datetime(df["Timestamp"], errors="coerce")
            df = df.dropna(subset=["Timestamp"])  # Remove invalid timestamps

            if df.empty:
                messagebox.showerror("Error", "No valid timestamps found in the DAT file.")
                return

            # Extract the month and year for sheet naming and headers
            first_date = df["Timestamp"].min()
            month_year = first_date.strftime("%B %Y")  # Example: "January 2025"
            sheet_name = f"DTR - {month_year}"

            # Filter for first time-in and last time-out
            final_df = filter_in_out_entries(df)

            if final_df.empty:
                messagebox.showerror("Error", "No data available for conversion. Check the input file.")
                return

            # Get save path from user
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile=os.path.basename(dat_file).replace(".dat", ".xlsx"),
                title="Save Converted Excel File"
            )

            if save_path:
                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    # Write the attendance sheet with a month-based name
                    final_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
                    
                    # Format the summary sheet to match the desired layout
                    format_dtr_summary_sheet(writer, final_df, month_year)

                    # Generate individual employee DTR sheets
                    for name in df["Name"].unique():
                        generate_employee_dtr(writer, df, name)

                # Auto-adjust column widths
                auto_adjust_column_widths(save_path)
                
                save_to_database(os.path.basename(dat_file), save_path)

                open_file = messagebox.askyesno("Conversion Complete", "File converted successfully!\nDo you want to open it now?")
                if open_file:
                    subprocess.run(["start", "", save_path], shell=True)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to convert file: {str(e)}")

def upload_employee_list():
    global employee_list
    file_path = filedialog.askopenfilename(
        title="Select Employee List TXT File",
        filetypes=[("Text Files", "*.txt")]
    )

    if file_path:
        try:
            conn = sqlite3.connect("conversion_history.db")
            cursor = conn.cursor()
            cursor.execute("DELETE FROM employees")  # Clear existing list

            with open(file_path, 'r') as file:
                employee_list.clear()
                for line in file:
                    parts = line.strip().split(' ', 1)
                    if len(parts) == 2:
                        emp_id = int(parts[0])
                        emp_name = parts[1].strip().upper()
                        employee_list[emp_id] = emp_name
                        cursor.execute("INSERT INTO employees (id, name) VALUES (?, ?)", (emp_id, emp_name))

            conn.commit()
            conn.close()
            employee_list_entry.insert(tk.END, file_path)  # Show file name in the listbox
            

        except Exception as e:
            messagebox.showerror("Error", f"Failed to upload employee list: {e}")

def load_employee_list():
    global employee_list
    conn = sqlite3.connect("conversion_history.db")
    cursor = conn.cursor()
    cursor.execute("SELECT id, name FROM employees")
    rows = cursor.fetchall()
    employee_list = {emp_id: name for emp_id, name in rows}
    conn.close()

def generate_employee_dtr(writer, df, employee_name):
    # Filter data for specific employee
    employee_df = df[df["Name"] == employee_name].copy()
    if employee_df.empty:
        return
    
    # Generate date range for the month
    first_date = employee_df["Timestamp"].min().date()
    year, month = first_date.year, first_date.month
    month_name = first_date.strftime('%B')
    first_day = datetime(year, month, 1)
    last_day = (first_day.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    all_dates = pd.date_range(first_day, last_day).date
    
    # Create the workbook
    worksheet = writer.book.create_sheet(employee_name)
    
    # --- Header Section ---
    # Create Civil Service form header
    worksheet.merge_cells('A1:E1')
    worksheet.merge_cells('I1:M1')
    cell = worksheet['A1']
    cell.value = "Civil Service Form No. 48"
    cell.alignment = Alignment(horizontal='center')
    
    cell = worksheet['I1']
    cell.value = "Civil Service Form No. 48"
    cell.alignment = Alignment(horizontal='center')
    
    # Create DAILY TIME RECORD header
    worksheet.merge_cells('A2:E2')
    worksheet.merge_cells('I2:M2')
    cell = worksheet['A2']
    cell.value = "DAILY TIME RECORD"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    
    cell = worksheet['I2']
    cell.value = "DAILY TIME RECORD"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    
    # Add Employee Name (bold, centered, and underlined) above "NAME"
    worksheet.merge_cells('A3:E3')
    worksheet.merge_cells('I3:M3')
    cell = worksheet['A3']
    cell.value = employee_name.upper()
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True, underline='single')  # Add underline
    
    cell = worksheet['I3']
    cell.value = employee_name.upper()
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True, underline='single')  # Add underline
    
    # Add NAME label
    worksheet.merge_cells('A4:E4')
    worksheet.merge_cells('I4:M4')
    cell = worksheet['A4']
    cell.value = "NAME"
    cell.alignment = Alignment(horizontal='center')
    
    cell = worksheet['I4']
    cell.value = "NAME"
    cell.alignment = Alignment(horizontal='center')
    
    # Add spacing row between NAME and "For the month of"
    worksheet.row_dimensions[5].height = 15  # Add space between NAME and month info
    
    # --- Month and Hours Section ---
    # Add month info (moved to row 6 instead of 5)
    month_range = f"{month_name} 1-{last_day.day}, {year}"
    
    worksheet.merge_cells('A6:E6')  # Changed from A5:E5 to A6:E6
    cell = worksheet['A6']  # Changed from A5 to A6
    cell.value = f"For the month of      {month_range}"
    
    worksheet.merge_cells('I6:M6')  # Changed from I5:M5 to I6:M6
    cell = worksheet['I6']  # Changed from I5 to I6
    cell.value = f"For the month of      {month_range}"
    
    # Add official hours info (adjust row numbers)
    worksheet.merge_cells('A7:E7')  # Changed from A6:E6 to A7:E7
    cell = worksheet['A7']  # Changed from A6 to A7
    cell.value = "Official hours for arrival and departure:"
    
    worksheet.merge_cells('I7:M7')  # Changed from I6:M6 to I7:M7
    cell = worksheet['I7']  # Changed from I6 to I7
    cell.value = "Official hours for arrival and departure:"
    
    # Extract and process time logs
    employee_df["Date"] = employee_df["Timestamp"].dt.date
    employee_df["Hour"] = employee_df["Timestamp"].dt.hour
    employee_df["Time"] = employee_df["Timestamp"].dt.strftime("%H:%M")
    
    # Create a calendar dictionary with default values
    calendar_data = {}
    
    # Fill in logs with default values
    for i, date in enumerate(all_dates, 1):
        date_str = date.strftime('%Y-%m-%d')
        weekday = date.weekday()
        
        # Default values
        if weekday == 5:  # Saturday
            calendar_data[i] = {
                'date': date,
                'arrival': '',
                'departure': '',
                'lunch_out': '',
                'lunch_in': '',
                'special': 'SATURDAY',
                'undertime': ''  # Add undertime field
            }
        elif weekday == 6:  # Sunday
            calendar_data[i] = {
                'date': date,
                'arrival': '',
                'departure': '',
                'lunch_out': '',
                'lunch_in': '',
                'special': 'SUNDAY',
                'undertime': ''  # Add undertime field
            }
        else:
            calendar_data[i] = {
                'date': date,
                'arrival': '',
                'departure': '',
                'lunch_out': '',
                'lunch_in': '',
                'special': 'ABSENT',  # Default to ABSENT for weekdays with no logs
                'undertime': ''  # Add undertime field
            }
    
    # Override with actual logs
    # Group logs by date
    logs_by_date = {}
    for _, row in employee_df.iterrows():
        date = row["Date"]
        day = date.day
        time = row["Time"]
        hour = row["Hour"]
        
        if day not in logs_by_date:
            logs_by_date[day] = []
        
        logs_by_date[day].append((time, hour))
    
    # Function to calculate undertime
    def calculate_undertime(arrival, departure, lunch_out, lunch_in):
        # Skip calculation if any time is missing
        if not arrival or not departure or not lunch_out or not lunch_in:
            return ""
        
        # If "No In" or "No Out" appears, return empty string
        if arrival == "No In" or departure == "No Out":
            return ""
        
        try:
            # Parse time strings to datetime objects
            arrival_time = datetime.strptime(arrival, "%H:%M")
            departure_time = datetime.strptime(departure, "%H:%M")
            lunch_out_time = datetime.strptime(lunch_out, "%H:%M")
            lunch_in_time = datetime.strptime(lunch_in, "%H:%M")
            
            # Standard work hours (9 hours including 1 hour lunch)
            std_arrival = datetime.strptime("08:00", "%H:%M")
            std_departure = datetime.strptime("17:00", "%H:%M")
            
            # Calculate actual work hours (excluding lunch)
            lunch_duration = (lunch_in_time - lunch_out_time).total_seconds() / 3600
            actual_work_hours = ((departure_time - arrival_time).total_seconds() / 3600) - lunch_duration
            
            # Standard work hours (excluding 1 hour lunch)
            standard_work_hours = 8
            
            # Calculate undertime in hours
            undertime_hours = max(0, standard_work_hours - actual_work_hours)
            
            # If late arrival or early departure
            if arrival_time > std_arrival:
                late_mins = (arrival_time - std_arrival).total_seconds() / 60
                undertime_hours += late_mins / 60
                
            if departure_time < std_departure:
                early_mins = (std_departure - departure_time).total_seconds() / 60
                undertime_hours += early_mins / 60
            
            # Format as HH:MM
            hours = int(undertime_hours)
            minutes = int((undertime_hours - hours) * 60)
            
            if hours > 0 or minutes > 0:
                return f"{hours:02}:{minutes:02}"
            else:
                return ""
        except:
            return ""
    
    # Process logs for each day
    for day, logs in logs_by_date.items():
        if day in calendar_data:
            # Clear the ABSENT special status if logs exist
            if calendar_data[day]['special'] == 'ABSENT':
                calendar_data[day]['special'] = ''
            
            # Sort by hour to get earliest and latest
            logs.sort(key=lambda x: x[1])
            
            if len(logs) == 1:
                # Single log - determine if it's arrival or departure based on time
                time, hour = logs[0]
                if hour < 12:  # Morning log
                    calendar_data[day]['arrival'] = time
                    calendar_data[day]['departure'] = "No Out"
                    # Set default lunch break for days with logs
                    calendar_data[day]['lunch_out'] = "12:01"
                    calendar_data[day]['lunch_in'] = "12:55"
                else:  # Afternoon log
                    calendar_data[day]['arrival'] = "No In"
                    calendar_data[day]['departure'] = time
                    # Set default lunch break for days with logs
                    calendar_data[day]['lunch_out'] = "12:01"
                    calendar_data[day]['lunch_in'] = "12:55"
            elif len(logs) >= 2:
                # Multiple logs - take first and last
                calendar_data[day]['arrival'] = logs[0][0]
                calendar_data[day]['departure'] = logs[-1][0]
                # Set default lunch break for days with logs
                calendar_data[day]['lunch_out'] = "12:01"
                calendar_data[day]['lunch_in'] = "12:55"
            
            # Calculate undertime (only for regular weekdays)
            if calendar_data[day]['special'] == '':
                calendar_data[day]['undertime'] = calculate_undertime(
                    calendar_data[day]['arrival'],
                    calendar_data[day]['departure'],
                    calendar_data[day]['lunch_out'],
                    calendar_data[day]['lunch_in']
                )
    
    # Count Saturdays with logs
    saturday_count = sum(1 for i, data in calendar_data.items() 
                         if data['date'].weekday() == 5 and 
                         (data['arrival'] or data['departure']))
    
    # Add regular days hours (adjust row numbers)
    worksheet.merge_cells('A8:E8')  # Changed from A7:E7 to A8:E8
    cell = worksheet['A8']  # Changed from A7 to A8
    cell.value = "Regular days: 8:00 AM - 5:00 PM"
    
    worksheet.merge_cells('I8:M8')  # Changed from I7:M7 to I8:M8
    cell = worksheet['I8']  # Changed from I7 to I8
    cell.value = "Regular days: 8:00 AM - 5:00 PM"
    
    # Add Saturdays info with count (adjust row numbers)
    worksheet.merge_cells('A9:E9')  # Changed from A8:E8 to A9:E9
    cell = worksheet['A9']  # Changed from A8 to A9
    cell.value = f"Saturdays: {saturday_count} day(s)"
    
    worksheet.merge_cells('I9:M9')  # Changed from I8:M8 to I9:M9
    cell = worksheet['I9']  # Changed from I8 to I9
    cell.value = f"Saturdays: {saturday_count} day(s)"
    
    # --- Table Headers --- (adjust row numbers)
    # Create table headers - First set
    worksheet['A10'] = "Day"  # Changed from A9 to A10
    worksheet['B10'] = "Arrival"  # Changed from B9 to B10
    worksheet['C10'] = "Departure"  # Changed from C9 to C10
    worksheet['D10'] = "Arrival"  # Changed from D9 to D10
    worksheet['E10'] = "Departure"  # Changed from E9 to E10
    worksheet['F10'] = "Undertime"  # Changed from F9 to F10
    
    # Create table headers - Second set
    worksheet['I10'] = "Day"  # Changed from I9 to I10
    worksheet['J10'] = "Arrival"  # Changed from J9 to J10
    worksheet['K10'] = "Departure"  # Changed from K9 to K10
    worksheet['L10'] = "Arrival"  # Changed from L9 to L10
    worksheet['M10'] = "Departure"  # Changed from M9 to M10
    worksheet['N10'] = "Undertime"  # Changed from N9 to N10
    
    # Style the headers
    for col in range(1, 7):
        col_letter = get_column_letter(col)
        cell = worksheet[f'{col_letter}10']  # Changed from row 9 to row 10
        cell.alignment = Alignment(horizontal='center')
        cell.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(border_style='thin'),
            right=openpyxl.styles.Side(border_style='thin'),
            top=openpyxl.styles.Side(border_style='thin'),
            bottom=openpyxl.styles.Side(border_style='thin')
        )
    
    for col in range(9, 15):
        col_letter = get_column_letter(col)
        cell = worksheet[f'{col_letter}10']  # Changed from row 9 to row 10
        cell.alignment = Alignment(horizontal='center')
        cell.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(border_style='thin'),
            right=openpyxl.styles.Side(border_style='thin'),
            top=openpyxl.styles.Side(border_style='thin'),
            bottom=openpyxl.styles.Side(border_style='thin')
        )
    
    # Fill the table with the calendar data
    row_start = 11  # Changed from 10 to 11
    total_undertime_mins = 0
    
    for day, data in calendar_data.items():
        if day <= 31:  # Ensure we don't go beyond the maximum days
            # Day number cell
            cell = worksheet[f'A{row_start}']
            cell.value = day
            cell.alignment = Alignment(horizontal='center')
            
            # For Saturday/Sunday, still show time if available
            if data['special'] in ['SATURDAY', 'SUNDAY']:
                if data['arrival'] or data['departure']:
                    # If they have logs, show them
                    cell = worksheet[f'B{row_start}']
                    cell.value = data['arrival'] if data['arrival'] else ''
                    cell.alignment = Alignment(horizontal='center')
                    
                    cell = worksheet[f'C{row_start}']
                    cell.value = data['lunch_out'] if data['lunch_out'] else ''
                    cell.alignment = Alignment(horizontal='center')
                    
                    cell = worksheet[f'D{row_start}']
                    cell.value = data['lunch_in'] if data['lunch_in'] else ''
                    cell.alignment = Alignment(horizontal='center')
                    
                    cell = worksheet[f'E{row_start}']
                    cell.value = data['departure'] if data['departure'] else ''
                    cell.alignment = Alignment(horizontal='center')
                    
                    cell = worksheet[f'F{row_start}']
                    cell.value = ''  # No undertime for weekends
                    cell.alignment = Alignment(horizontal='center')
                else:
                    # No logs, just show the day type
                    worksheet.merge_cells(f'B{row_start}:F{row_start}')
                    cell = worksheet[f'B{row_start}']
                    cell.value = data['special']
                    cell.alignment = Alignment(horizontal='center')
            elif data['special'] == 'ABSENT':
                # For absent days
                worksheet.merge_cells(f'B{row_start}:F{row_start}')
                cell = worksheet[f'B{row_start}']
                cell.value = 'ABSENT'
                cell.alignment = Alignment(horizontal='center')
            else:
                # Regular day with time entries
                cell = worksheet[f'B{row_start}']
                cell.value = data['arrival']
                cell.alignment = Alignment(horizontal='center')
                
                cell = worksheet[f'C{row_start}']
                cell.value = data['lunch_out']
                cell.alignment = Alignment(horizontal='center')
                
                cell = worksheet[f'D{row_start}']
                cell.value = data['lunch_in']
                cell.alignment = Alignment(horizontal='center')
                
                cell = worksheet[f'E{row_start}']
                cell.value = data['departure']
                cell.alignment = Alignment(horizontal='center')
                
                cell = worksheet[f'F{row_start}']
                cell.value = data['undertime']  # Add the calculated undertime
                cell.alignment = Alignment(horizontal='center')
                
                # Track total undertime
                if data['undertime']:
                    try:
                        hrs, mins = data['undertime'].split(':')
                        total_undertime_mins += int(hrs) * 60 + int(mins)
                    except:
                        pass
            
            # Add borders to cells
            for col in range(1, 7):
                col_letter = get_column_letter(col)
                worksheet[f'{col_letter}{row_start}'].border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(border_style='thin'),
                    right=openpyxl.styles.Side(border_style='thin'),
                    top=openpyxl.styles.Side(border_style='thin'),
                    bottom=openpyxl.styles.Side(border_style='thin')
                )
            
            # Repeat for the second half (identical data for demo purposes)
            # Day number cell for second half
            cell = worksheet[f'I{row_start}']
            cell.value = day
            cell.alignment = Alignment(horizontal='center')
            
            # For Saturday/Sunday, still show time if available
            if data['special'] in ['SATURDAY', 'SUNDAY']:
                if data['arrival'] or data['departure']:
                    # If they have logs, show them
                    cell = worksheet[f'J{row_start}']
                    cell.value = data['arrival'] if data['arrival'] else ''
                    cell.alignment = Alignment(horizontal='center')
                    
                    cell = worksheet[f'K{row_start}']
                    cell.value = data['lunch_out'] if data['lunch_out'] else ''
                    cell.alignment = Alignment(horizontal='center')
                    
                    cell = worksheet[f'L{row_start}']
                    cell.value = data['lunch_in'] if data['lunch_in'] else ''
                    cell.alignment = Alignment(horizontal='center')
                    
                    cell = worksheet[f'M{row_start}']
                    cell.value = data['departure'] if data['departure'] else ''
                    cell.alignment = Alignment(horizontal='center')
                    
                    cell = worksheet[f'N{row_start}']
                    cell.value = ''  # No undertime for weekends
                    cell.alignment = Alignment(horizontal='center')
                else:
                    # No logs, just show the day type
                    worksheet.merge_cells(f'J{row_start}:N{row_start}')
                    cell = worksheet[f'J{row_start}']
                    cell.value = data['special']
                    cell.alignment = Alignment(horizontal='center')
            elif data['special'] == 'ABSENT':
                # For absent days
                worksheet.merge_cells(f'J{row_start}:N{row_start}')
                cell = worksheet[f'J{row_start}']
                cell.value = 'ABSENT'
                cell.alignment = Alignment(horizontal='center')
            else:
                # Regular day with time entries
                cell = worksheet[f'J{row_start}']
                cell.value = data['arrival']
                cell.alignment = Alignment(horizontal='center')
                
                cell = worksheet[f'K{row_start}']
                cell.value = data['lunch_out']
                cell.alignment = Alignment(horizontal='center')
                
                cell = worksheet[f'L{row_start}']
                cell.value = data['lunch_in']
                cell.alignment = Alignment(horizontal='center')
                
                cell = worksheet[f'M{row_start}']
                cell.value = data['departure']
                cell.alignment = Alignment(horizontal='center')
                
                cell = worksheet[f'N{row_start}']
                cell.value = data['undertime']  # Add the calculated undertime
                cell.alignment = Alignment(horizontal='center')
            
            # Add borders to cells in second half
            for col in range(9, 15):
                col_letter = get_column_letter(col)
                worksheet[f'{col_letter}{row_start}'].border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(border_style='thin'),
                    right=openpyxl.styles.Side(border_style='thin'),
                    top=openpyxl.styles.Side(border_style='thin'),
                    bottom=openpyxl.styles.Side(border_style='thin')
                )
                
            row_start += 1
    
    # Convert total undertime minutes to hours and minutes
    total_undertime_hours = total_undertime_mins // 60
    total_undertime_minutes = total_undertime_mins % 60
    total_undertime_str = f"{total_undertime_hours:02}:{total_undertime_minutes:02}"
    
    # Add Total row
    cell = worksheet[f'A{row_start}']
    cell.value = "Total"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    
    # Add total undertime
    cell = worksheet[f'F{row_start}']
    cell.value = total_undertime_str if total_undertime_mins > 0 else ""
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    
    # Add borders to total row
    for col in range(1, 7):
        col_letter = get_column_letter(col)
        worksheet[f'{col_letter}{row_start}'].border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(border_style='thin'),
            right=openpyxl.styles.Side(border_style='thin'),
            top=openpyxl.styles.Side(border_style='thin'),
            bottom=openpyxl.styles.Side(border_style='thin')
        )
    
    # Repeat for second half
    cell = worksheet[f'I{row_start}']
    cell.value = "Total"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    
    # Add total undertime to second half
    cell = worksheet[f'N{row_start}']
    cell.value = total_undertime_str if total_undertime_mins > 0 else ""
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    
    # Add borders to total row in second half
    for col in range(9, 15):
        col_letter = get_column_letter(col)
        worksheet[f'{col_letter}{row_start}'].border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(border_style='thin'),
            right=openpyxl.styles.Side(border_style='thin'),
            top=openpyxl.styles.Side(border_style='thin'),
            bottom=openpyxl.styles.Side(border_style='thin')
        )
    
    # Add certification text
    certification_row = row_start + 2
    
    # First DTR certification
    worksheet.merge_cells(f'A{certification_row}:F{certification_row}')
    cell = worksheet[f'A{certification_row}']
    cell.value = "I certify on my honor that the above is a true and"
    cell.alignment = Alignment(horizontal='center')
    
    cert_row2 = certification_row + 1
    worksheet.merge_cells(f'A{cert_row2}:F{cert_row2}')
    cell = worksheet[f'A{cert_row2}']
    cell.value = "correct report of the hours of work performed, record"
    cell.alignment = Alignment(horizontal='center')
    
    cert_row3 = cert_row2 + 1
    worksheet.merge_cells(f'A{cert_row3}:F{cert_row3}')
    cell = worksheet[f'A{cert_row3}']
    cell.value = "of which was made daily at the time of arrival and"
    cell.alignment = Alignment(horizontal='center')
    
    cert_row4 = cert_row3 + 1
    worksheet.merge_cells(f'A{cert_row4}:F{cert_row4}')
    cell = worksheet[f'A{cert_row4}']
    cell.value = "departure from office."
    cell.alignment = Alignment(horizontal='center')
    
    # Second DTR certification (mirror of the first)
    worksheet.merge_cells(f'I{certification_row}:N{certification_row}')
    cell = worksheet[f'I{certification_row}']
    cell.value = "I certify on my honor that the above is a true and"
    cell.alignment = Alignment(horizontal='center')
    
    worksheet.merge_cells(f'I{cert_row2}:N{cert_row2}')
    cell = worksheet[f'I{cert_row2}']
    cell.value = "correct report of the hours of work performed, record"
    cell.alignment = Alignment(horizontal='center')
    
    worksheet.merge_cells(f'I{cert_row3}:N{cert_row3}')
    cell = worksheet[f'I{cert_row3}']
    cell.value = "of which was made daily at the time of arrival and"
    cell.alignment = Alignment(horizontal='center')
    
    worksheet.merge_cells(f'I{cert_row4}:N{cert_row4}')
    cell = worksheet[f'I{cert_row4}']
    cell.value = "departure from office."
    cell.alignment = Alignment(horizontal='center')
    
    # Add signature lines
    sig_row = cert_row4 + 3  # Add space before signature line
    
    # First DTR signature line
    worksheet.merge_cells(f'A{sig_row}:F{sig_row}')
    cell = worksheet[f'A{sig_row}']
    cell.value = "_" * 40
    cell.alignment = Alignment(horizontal='center')
    
    sig_label_row = sig_row + 1
    worksheet.merge_cells(f'A{sig_label_row}:F{sig_label_row}')
    cell = worksheet[f'A{sig_label_row}']
    cell.value = "Signature of Employee"
    cell.alignment = Alignment(horizontal='center')
    
    # Second DTR signature line
    worksheet.merge_cells(f'I{sig_row}:N{sig_row}')
    cell = worksheet[f'I{sig_row}']
    cell.value = "_" * 40
    cell.alignment = Alignment(horizontal='center')
    
    worksheet.merge_cells(f'I{sig_label_row}:N{sig_label_row}')
    cell = worksheet[f'I{sig_label_row}']
    cell.value = "Signature of Employee"
    cell.alignment = Alignment(horizontal='center')
    
    # Add verification text
    verify_row = sig_label_row + 2
    
    # First DTR verification
    worksheet.merge_cells(f'A{verify_row}:F{verify_row}')
    cell = worksheet[f'A{verify_row}']
    cell.value = "Verified as to the prescribed office hours."
    cell.alignment = Alignment(horizontal='center')
    
    # Second DTR verification
    worksheet.merge_cells(f'I{verify_row}:N{verify_row}')
    cell = worksheet[f'I{verify_row}']
    cell.value = "Verified as to the prescribed office hours."
    cell.alignment = Alignment(horizontal='center')
    
    # Add supervisor signature lines
    super_sig_row = verify_row + 3
    
    # First DTR supervisor signature
    worksheet.merge_cells(f'A{super_sig_row}:F{super_sig_row}')
    cell = worksheet[f'A{super_sig_row}']
    cell.value = "_" * 40
    cell.alignment = Alignment(horizontal='center')
    
    # Second DTR supervisor signature
    worksheet.merge_cells(f'I{super_sig_row}:N{super_sig_row}')
    cell = worksheet[f'I{super_sig_row}']
    cell.value = "_" * 40
    cell.alignment = Alignment(horizontal='center')
    
    # Add supervisor name and position
    name_row = super_sig_row + 1
    
    # First DTR supervisor name
    worksheet.merge_cells(f'A{name_row}:F{name_row}')
    cell = worksheet[f'A{name_row}']
    cell.value = "FORTUNATO L. PALILEO"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    
    # Second DTR supervisor name
    worksheet.merge_cells(f'I{name_row}:N{name_row}')
    cell = worksheet[f'I{name_row}']
    cell.value = "FORTUNATO L. PALILEO"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    
    # Add supervisor position
    pos_row = name_row + 1
    
    # First DTR supervisor position
    worksheet.merge_cells(f'A{pos_row}:F{pos_row}')
    cell = worksheet[f'A{pos_row}']
    cell.value = "CHIEF, EDP SERVICES"
    cell.alignment = Alignment(horizontal='center')
    
    # Second DTR supervisor position
    worksheet.merge_cells(f'I{pos_row}:N{pos_row}')
    cell = worksheet[f'I{pos_row}']
    cell.value = "CHIEF, EDP SERVICES"
    cell.alignment = Alignment(horizontal='center')
    
    # Set column widths to match expected format (this stays as it was)
    worksheet.column_dimensions['A'].width = 8
    worksheet.column_dimensions['B'].width = 8
    worksheet.column_dimensions['C'].width = 8
    worksheet.column_dimensions['D'].width = 8
    worksheet.column_dimensions['E'].width = 8
    worksheet.column_dimensions['F'].width = 10
    
    # Add gap columns (G, H)
    worksheet.column_dimensions['G'].width = 5
    worksheet.column_dimensions['H'].width = 5
    
    # Second DTR section columns
    worksheet.column_dimensions['I'].width = 8
    worksheet.column_dimensions['J'].width = 8
    worksheet.column_dimensions['K'].width = 8
    worksheet.column_dimensions['L'].width = 8
    worksheet.column_dimensions['M'].width = 8
    worksheet.column_dimensions['N'].width = 10

from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def auto_adjust_column_widths(file_path):
    wb = load_workbook(file_path)
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        column_widths = {}

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # Check if the cell is part of a merged range
                    if any(cell.coordinate in merged_range for merged_range in ws.merged_cells):
                        continue  # Skip merged cells
                    
                    col_letter = get_column_letter(cell.column)
                    column_widths[col_letter] = max(column_widths.get(col_letter, 0), len(str(cell.value)))

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width + 2  # Add padding

    wb.save(file_path)

def show_history():
    global history_window, history_window_open

    if not history_window_open:
        history_window_open = True

        def filter_history():
            search_query = search_entry.get().lower()
            for item in tree.get_children():
                tree.delete(item)
            conn = sqlite3.connect("conversion_history.db")
            cursor = conn.cursor()
            cursor.execute("SELECT filename, converted_at, output_path FROM conversions ORDER BY id DESC")
            records = cursor.fetchall()
            conn.close()
            for record in records:
                if search_query in record[0].lower() or search_query in record[1]:
                    tree.insert("", "end", values=record)

        history_window = tk.Toplevel(root)
        history_window.title("Conversion History")
        history_window.geometry("550x350")
        history_window.configure(bg="#f5f5f5")

        search_frame = tk.Frame(history_window, bg="#f5f5f5")
        search_frame.pack(pady=5)
        search_label = tk.Label(search_frame, text="Search:", bg="#f5f5f5")
        search_label.pack(side=tk.LEFT, padx=5)
        search_entry = tk.Entry(search_frame)
        search_entry.pack(side=tk.LEFT, padx=5)
        search_button = tk.Button(search_frame, text="Filter", command=filter_history)
        search_button.pack(side=tk.LEFT, padx=5)

        tree = ttk.Treeview(history_window, columns=("Filename", "Date", "Output Path"), show="headings")
        tree.heading("Filename", text="Filename")
        tree.heading("Date", text="Date Converted")
        tree.heading("Output Path", text="Output Path")
        tree.column("Filename", width=150)
        tree.column("Date", width=120)
        tree.column("Output Path", width=200)
        tree.pack(expand=True, fill="both")
        filter_history()

        history_window.protocol("WM_DELETE_WINDOW", close_history_window)
    else:
        messagebox.showinfo("Info", "History window is already open.")

def close_history_window():
    global history_window, history_window_open
    if history_window:
        history_window.destroy()
        history_window = None
    history_window_open = False

def browse_dat_files():
    file_paths = filedialog.askopenfilenames(filetypes=[("Data Files", "*.dat")])
    if not file_paths:
        return

    for file_path in file_paths:
        employee_list_entry.insert(tk.END, file_path)  # Add filename to the DAT listbox

def upload_employee_list():
    global employee_list
    file_path = filedialog.askopenfilename(
        title="Select Employee List TXT File",
        filetypes=[("Text Files", "*.txt")]
    )

    if file_path:
        try:
            conn = sqlite3.connect("conversion_history.db")
            cursor = conn.cursor()
            cursor.execute("DELETE FROM employees")  # Clear existing list

            with open(file_path, 'r') as file:
                employee_list.clear()
                for line in file:
                    parts = line.strip().split(' ', 1)
                    if len(parts) == 2:
                        emp_id = int(parts[0])
                        emp_name = parts[1].strip().upper()
                        employee_list[emp_id] = emp_name
                        cursor.execute("INSERT INTO employees (id, name) VALUES (?, ?)", (emp_id, emp_name))

            conn.commit()
            conn.close()
            employee_list_entry.insert(tk.END, file_path)  # Show file name in the TXT listbox
            

        except Exception as e:
            messagebox.showerror("Error", f"Failed to upload employee list: {e}")

def preview_dat_file(event):
    selected_index = employee_list_entry.curselection()
    if not selected_index:
        return

    file_path = employee_list_entry.get(selected_index[0])

    try:
        df = pd.read_csv(file_path, delimiter="\t", dtype=str)  # Read data file

        preview_window = tk.Toplevel(root)
        preview_window.title(f"Preview: {os.path.basename(file_path)}")
        preview_window.geometry("800x500")

        frame = tk.Frame(preview_window)
        frame.pack(expand=True, fill="both", padx=10, pady=10)

        tree = ttk.Treeview(frame, show="headings")

        # Set up columns
        tree["columns"] = list(df.columns)
        for col in df.columns:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=max(df[col].astype(str).str.len().max() * 8, 100))  # Auto adjust width

        # Insert rows
        for _, row in df.iterrows():
            tree.insert("", "end", values=list(row))

        # Add scrollbar
        scrollbar_x = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        scrollbar_y = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(xscrollcommand=scrollbar_x.set, yscrollcommand=scrollbar_y.set)

        scrollbar_x.pack(side="bottom", fill="x")
        scrollbar_y.pack(side="right", fill="y")
        tree.pack(expand=True, fill="both")

        # Add Convert button for DAT files
        def proceed_to_conversion():
            preview_window.destroy()
            convert_batch_to_excel([file_path])

        proceed_button = tk.Button(preview_window, text="Convert", command=proceed_to_conversion, 
                                  font=("Segoe UI", 12, "bold"), fg="white", bg="#4CAF50", 
                                  relief="flat", padx=10, pady=5)
        proceed_button.pack(pady=10)

    except Exception as e:
        messagebox.showerror("Error", f"Failed to load file: {e}")

def preview_txt_file(event):
    selected_index = employee_list_entry.curselection()
    if not selected_index:
        return

    file_path = employee_list_entry.get(selected_index[0])

    try:
        with open(file_path, 'r') as file:
            content = file.readlines()

        preview_window = tk.Toplevel(root)
        preview_window.title(f"Preview: {os.path.basename(file_path)}")
        preview_window.geometry("600x400")

        frame = tk.Frame(preview_window)
        frame.pack(expand=True, fill="both", padx=10, pady=10)

        # Create a Text widget to display the content
        text_widget = tk.Text(frame, wrap="none", font=("Courier New", 12))
        
        # Add scrollbars
        scrollbar_y = tk.Scrollbar(frame, orient="vertical", command=text_widget.yview)
        scrollbar_x = tk.Scrollbar(frame, orient="horizontal", command=text_widget.xview)
        text_widget.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        # Insert content
        for line in content:
            text_widget.insert(tk.END, line)
        
        # Make it read-only
        text_widget.config(state="disabled")
        
        # Arrange widgets
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
        text_widget.pack(side="left", fill="both", expand=True)

    except Exception as e:
        messagebox.showerror("Error", f"Failed to load file: {e}")

def save_to_database(filename, output_path):
    conn = sqlite3.connect("conversion_history.db")
    cursor = conn.cursor()
    cursor.execute("INSERT INTO conversions (filename, converted_at, output_path) VALUES (?, ?, ?)", 
                   (filename, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), output_path))
    conn.commit()
    conn.close()

def load_excel_history():
    conn = sqlite3.connect("conversion_history.db")
    cursor = conn.cursor()
    cursor.execute("SELECT output_path FROM conversions ORDER BY id DESC")
    records = cursor.fetchall()
    conn.close()
    
    # Add existing Excel files to the listbox
    for record in records:
        if os.path.exists(record[0]):  # Only add if the file still exists
            employee_list_entry.insert(tk.END, record[0])

def open_excel_file(event):
    selected_index = employee_list_entry.curselection()
    if not selected_index:
        return

    file_path = employee_list_entry.get(selected_index[0])
    
    if os.path.exists(file_path):
        try:
            subprocess.run(["start", "", file_path], shell=True)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open file: {e}")
    else:
        messagebox.showerror("Error", "File not found. It may have been moved or deleted.")
        # Remove entry from listbox if file doesn't exist
        employee_list_entry.delete(selected_index)

def browse_and_preview_employee_list():
    file_path = filedialog.askopenfilename(
        title="Select Employee List TXT File",
        filetypes=[("Text Files", "*.txt")]
    )

    if file_path:
        # Clear previous entries
        employee_list_entry.delete(0, tk.END)
        employee_list_entry.insert(0, file_path)

        preview_employee_list.config(state=tk.NORMAL)
        preview_employee_list.delete(1.0, tk.END)
        
        try:
            with open(file_path, 'r') as file:
                content = file.read()
                preview_employee_list.insert(tk.END, content)
        except Exception as e:
            preview_employee_list.insert(tk.END, f"Error reading file: {e}")
        
        preview_employee_list.config(state=tk.DISABLED)

def browse_and_preview_dat_files():
    files = filedialog.askopenfilenames(
        title="Select DAT Files",
        filetypes=[("DAT Files", "*.dat")]
    )

    if files:
        dat_file_entry.delete(0, tk.END)
        dat_file_entry.insert(0, ", ".join(files))

        preview_dat_files.config(state=tk.NORMAL)
        preview_dat_files.delete(1.0, tk.END)
        
        for file_path in files:
            try:
                with open(file_path, 'r') as file:
                    content = file.read()
                    preview_dat_files.insert(tk.END, f"File: {os.path.basename(file_path)}\n")
                    preview_dat_files.insert(tk.END, content + "\n\n")
            except Exception as e:
                preview_dat_files.insert(tk.END, f"Error reading {file_path}: {e}\n\n")
        
        preview_dat_files.config(state=tk.DISABLED)
def convert_files():
    employee_file = employee_list_entry.get()
    dat_files = dat_file_entry.get().split(", ")

    if not employee_file or not dat_files:
        messagebox.showwarning("Warning", "Please upload both Employee List and DAT Files")
        return

    try:
        # Update employee list first
        upload_employee_list_from_path(employee_file)

        # Convert DAT files
        convert_batch_to_excel(dat_files)

    except Exception as e:
        messagebox.showerror("Error", str(e))

def upload_employee_list_from_path(file_path):
    global employee_list
    try:
        conn = sqlite3.connect("conversion_history.db")
        cursor = conn.cursor()
        cursor.execute("DELETE FROM employees")  # Clear existing list

        with open(file_path, 'r') as file:
            employee_list.clear()
            for line in file:
                parts = line.strip().split(' ', 1)
                if len(parts) == 2:
                    emp_id = int(parts[0])
                    emp_name = parts[1].strip().upper()
                    employee_list[emp_id] = emp_name
                    cursor.execute("INSERT INTO employees (id, name) VALUES (?, ?)", (emp_id, emp_name))

        conn.commit()
        conn.close()
        

    except Exception as e:
        messagebox.showerror("Error", f"Failed to upload employee list: {e}")

class StyledTkinter:
    # Color Palette
    COLORS = {
        'bg_primary': '#f8f9fa',      # Light gray-white
        'bg_secondary': '#f1f3f5',    # Slightly darker white
        'bg_accent': '#e9ecef',       # Light gray accent

        'text_primary': '#1a365d',    # Dark navy blue
        'text_secondary': '#2c3e50',  # Slightly lighter navy blue
        'btn_primary': '#3182ce',     # Vibrant blue
        'btn_secondary': '#4a5568',   # Dark grayish blue
        'btn_success': '#48bb78',     # Green for success
        'btn_warning': '#ed8936',     # Orange for warnings
        'border_color': '#cbd5e0'     # Light border color
    }

    @classmethod
    def create_styled_button(cls, parent, text, command, style='primary', width=None):
        btn_styles = {
            'primary': {
                'bg': cls.COLORS['btn_primary'], 
                'fg': 'white', 
                'hover_bg': '#2c5282'
            },
            'secondary': {
                'bg': cls.COLORS['btn_secondary'], 
                'fg': 'white', 
                'hover_bg': '#718096'
            },
            'success': {
                'bg': cls.COLORS['btn_success'], 
                'fg': 'white', 
                'hover_bg': '#38a169'
            },
            'warning': {
                'bg': cls.COLORS['btn_warning'], 
                'fg': 'white', 
                'hover_bg': '#dd6b20'
            }
        }

        current_style = btn_styles.get(style, btn_styles['primary'])
        
        btn = tk.Button(
            parent, 
            text=text, 
            command=command,
            bg=current_style['bg'], 
            fg=current_style['fg'],
            font=("Segoe UI", 10, "bold"),
            relief=tk.FLAT,
            padx=10,
            pady=5
        )

        if width:
            btn.config(width=width)

        # Hover effects
        def on_enter(e):
            btn.config(bg=current_style['hover_bg'])

        def on_leave(e):
            btn.config(bg=current_style['bg'])

        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)

        return btn

    @classmethod
    def create_styled_entry(cls, parent, width=50):
        entry = tk.Entry(
            parent, 
            font=("Segoe UI", 10), 
            width=width,
            bg=cls.COLORS['bg_secondary'],
            fg=cls.COLORS['text_primary'],
            insertbackground=cls.COLORS['text_primary'],
            relief=tk.FLAT,
            highlightthickness=1,
            highlightcolor=cls.COLORS['border_color'],
            highlightbackground=cls.COLORS['border_color']
        )
        return entry

    @classmethod
    def create_styled_label(cls, parent, text, style='primary'):
        label_styles = {
            'primary': {
                'fg': cls.COLORS['text_primary'],
                'font': ("Segoe UI", 10, "bold")
            },
            'secondary': {
                'fg': cls.COLORS['text_secondary'],
                'font': ("Segoe UI", 10)
            }
        }
        
        current_style = label_styles.get(style, label_styles['primary'])
        
        label = tk.Label(
            parent, 
            text=text, 
            fg=current_style['fg'],
            font=current_style['font'],
            bg=cls.COLORS['bg_primary']
        )
        return label

def create_improved_gui():
    root = tk.Tk()
    root.title("Employee DTR Converter")
    root.geometry("800x800")
    root.configure(bg=StyledTkinter.COLORS['bg_primary'])

    # Main Container
    main_container = tk.Frame(root, bg=StyledTkinter.COLORS['bg_primary'], padx=30, pady=30)
    main_container.pack(fill=tk.BOTH, expand=True)

    # Title
    title_label = StyledTkinter.create_styled_label(
        main_container, 
        "Employee DTR Converter", 
        style='primary'
    )
    title_label.config(font=("Segoe UI", 18, "bold"))
    title_label.pack(pady=(0, 20))

    # Employee List Section
    employee_section = tk.LabelFrame(
        main_container, 
        text="Employee List", 
        font=("Segoe UI", 12, "bold"), 
        bg=StyledTkinter.COLORS['bg_primary'],
        fg=StyledTkinter.COLORS['text_primary'],
        labelanchor='n',
        borderwidth=2,
        relief=tk.GROOVE
    )
    employee_section.pack(fill=tk.X, pady=10)

    employee_row = tk.Frame(employee_section, bg=StyledTkinter.COLORS['bg_primary'])
    employee_row.pack(padx=10, pady=10, fill=tk.X)

    # Employee List Label
    employee_list_label = StyledTkinter.create_styled_label(
        employee_row, 
        "Employee List File:", 
        style='secondary'
    )
    employee_list_label.pack(side=tk.LEFT, padx=(0, 10))

    # Employee List Entry
    employee_list_entry = StyledTkinter.create_styled_entry(employee_row, width=50)
    employee_list_entry.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 10))

    # Employee List Browse Button
    employee_upload_btn = StyledTkinter.create_styled_button(
        employee_row, 
        "Browse", 
        lambda: browse_and_preview_employee_list(),
        style='primary'
    )
    employee_upload_btn.pack(side=tk.LEFT)

    # Employee List Preview Frame
    employee_preview_frame = tk.Frame(employee_section, bg=StyledTkinter.COLORS['bg_primary'])
    employee_preview_frame.pack(padx=10, pady=10, fill=tk.X)

    # Employee List Preview
    preview_employee_list = tk.Text(
        employee_preview_frame, 
        height=10, 
        font=("Courier", 10), 
        state=tk.DISABLED, 
        wrap=tk.NONE,
        bg=StyledTkinter.COLORS['bg_secondary'],
        fg=StyledTkinter.COLORS['text_primary']
    )
    preview_employee_list.pack(side=tk.LEFT, expand=True, fill=tk.X)

    # Employee Preview Scrollbars
    employee_preview_scrollbar_y = tk.Scrollbar(
        employee_preview_frame, 
        orient=tk.VERTICAL, 
        command=preview_employee_list.yview,
        bg=StyledTkinter.COLORS['bg_accent']
    )
    employee_preview_scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
    
    employee_preview_scrollbar_x = tk.Scrollbar(
        employee_section, 
        orient=tk.HORIZONTAL, 
        command=preview_employee_list.xview,
        bg=StyledTkinter.COLORS['bg_accent']
    )
    employee_preview_scrollbar_x.pack(fill=tk.X, padx=10)
    
    preview_employee_list.configure(
        yscrollcommand=employee_preview_scrollbar_y.set,
        xscrollcommand=employee_preview_scrollbar_x.set
    )

    # DAT Files Section
    dat_section = tk.LabelFrame(
        main_container, 
        text="DTR Files", 
        font=("Segoe UI", 12, "bold"), 
        bg=StyledTkinter.COLORS['bg_primary'],
        fg=StyledTkinter.COLORS['text_primary'],
        labelanchor='n',
        borderwidth=2,
        relief=tk.GROOVE
    )
    dat_section.pack(fill=tk.X, pady=10)

    dat_row = tk.Frame(dat_section, bg=StyledTkinter.COLORS['bg_primary'])
    dat_row.pack(padx=10, pady=10, fill=tk.X)

    # DAT Files Label
    dat_file_label = StyledTkinter.create_styled_label(
        dat_row, 
        "DTR Files:", 
        style='secondary'
    )
    dat_file_label.pack(side=tk.LEFT, padx=(0, 10))

    # DAT Files Entry
    dat_file_entry = StyledTkinter.create_styled_entry(dat_row, width=50)
    dat_file_entry.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 10))

    # DAT Files Browse Button
    dat_upload_btn = StyledTkinter.create_styled_button(
        dat_row, 
        "Browse", 
        lambda: browse_and_preview_dat_files(),
        style='primary'
    )
    dat_upload_btn.pack(side=tk.LEFT)

    # DAT Files Preview Frame
    dat_preview_frame = tk.Frame(dat_section, bg=StyledTkinter.COLORS['bg_primary'])
    dat_preview_frame.pack(padx=10, pady=10, fill=tk.X)

    # DAT Files Preview
    preview_dat_files = tk.Text(
        dat_preview_frame, 
        height=10, 
        font=("Courier", 10), 
        state=tk.DISABLED, 
        wrap=tk.NONE,
        bg=StyledTkinter.COLORS['bg_secondary'],
        fg=StyledTkinter.COLORS['text_primary']
    )
    preview_dat_files.pack(side=tk.LEFT, expand=True, fill=tk.X)

    # DAT Preview Scrollbars
    dat_preview_scrollbar_y = tk.Scrollbar(
        dat_preview_frame, 
        orient=tk.VERTICAL, 
        command=preview_dat_files.yview,
        bg=StyledTkinter.COLORS['bg_accent']
    )
    dat_preview_scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
    
    dat_preview_scrollbar_x = tk.Scrollbar(
        dat_section, 
        orient=tk.HORIZONTAL, 
        command=preview_dat_files.xview,
        bg=StyledTkinter.COLORS['bg_accent']
    )
    dat_preview_scrollbar_x.pack(fill=tk.X, padx=10)
    
    preview_dat_files.configure(
        yscrollcommand=dat_preview_scrollbar_y.set,
        xscrollcommand=dat_preview_scrollbar_x.set
    )

    # Conversion and History Buttons
    button_frame = tk.Frame(main_container, bg=StyledTkinter.COLORS['bg_primary'])
    button_frame.pack(pady=10)

    history_btn = StyledTkinter.create_styled_button(
        button_frame, 
        "View Conversion History", 
        show_history,
        style='secondary'
    )
    history_btn.pack(side=tk.LEFT, padx=10)

    convert_btn = StyledTkinter.create_styled_button(
        button_frame, 
        "Convert DTR to Excel", 
        convert_files,
        style='success'
    )
    convert_btn.pack(side=tk.LEFT, padx=10)

    return root, employee_list_entry, preview_employee_list, dat_file_entry, preview_dat_files
    

# Replace the original create_gui() function with this new one
# Modify the main block
if __name__ == "__main__":
    create_database()
    load_employee_list()  # Load employee list from DB if any
    
    # Create the root window and key widgets using the new method
    root, employee_list_entry, preview_employee_list, dat_file_entry, preview_dat_files = create_improved_gui()
    
    # Update global references if needed
    globals()['employee_list_entry'] = employee_list_entry
    globals()['preview_employee_list'] = preview_employee_list
    globals()['dat_file_entry'] = dat_file_entry
    globals()['preview_dat_files'] = preview_dat_files
    
    root.mainloop()